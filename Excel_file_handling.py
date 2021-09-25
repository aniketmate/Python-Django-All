# import openpyxl as xcl
# from openpyxl.cell import cell

# wb = openpyxl.Workbook()
# # print(wb)
# sheet = wb.active
# print(sheet)


# sheet["A1"] = "ABC"
# sheet["B1"] = "PQR"
# sheet["C1"] = "XYZ"


# sheet.cell(1,1).value = "A"
# sheet.cell(1,2).value = "B"
# sheet.cell(1,3).value = "C"
# sheet.cell(1,4).value = "D"

# sheet.cell(1,1, "P")
# sheet.cell(1,2, "Q")
# sheet.cell(1,3, "R")
# sheet.cell(1,4, "S")


# import time
# now = time.strftime("%H %M %S")
# sheet.cell(1,1, value=now)

# try:
#     wb.save("E:\\aniket mate\\Excel_file_handling\\first.xlsx")
# except PermissionError:
#     print("Plz close the file ")
#     print("Plz close file...!")

# FILE_PATH = "E:\\aniket mate\\Excel_file_handling\\first.xlsx"

# wb = xcl.load_workbook(FILE_PATH)
# sheet = wb.active
# print(sheet.min_row)
# print(sheet.min_column)
# print(sheet.max_row)
# print(sheet.max_column)
# sheet.merge_cells("A9:C9")
# print(sheet.cell(5,3).value)

# for i in range(4,90):
#     print(sheet.cell(row=i, column=3).value)

# sheet.cell(row=91,column=3).value="Aniket"
# wb.save(FILE_PATH)

# data = (  
#     (11, 48, 50),  
#     (81, 30, 82),  
#     (20, 51, 72),  
#     (21, 14, 60),  
#     (28, 41, 49),  
#     (74, 65, 53),  
#     ("Peter", 'Andrew',45.63)  
# )  

# for i in data:
#     sheet.append(i)
# c = 1 
# for i in data:
#     sheet.cell(c, 1).value = i[0]
#     sheet.cell(c, 2).value = i[1]
#     sheet.cell(c, 3).value = i[2]
#     c += 1

# data = sheet["A1":"C7"]
# for f,s,t in data:
#     print(f.value,s.value,t.value)

# for row in sheet.iter_rows(min_row=sheet.min_row, min_col=sheet.min_column, max_row=sheet.max_row, max_col=sheet.max_column):  
#     print(row)  # (<Cell 'Sheet'.A1>, <Cell 'Sheet'.B1>, <Cell 'Sheet'.C1>)
#     for cell in row:
#         print(cell.value, end=' ')
#     print()

# for row in sheet.iter_cols(min_row=1, min_col=1, max_row=7, max_col=3):  
#     print(row)
#     for cell in row:  
#         print(cell.value, end=" ")  
    # print()  
# wb.save(FILE_PATH)

import openpyxl
wb = openpyxl.Workbook()
# print(wb)
sheet = wb.active
# print(sheet)
# for i in range (1,51):
#     for j in range(1,51):
#         sheet.cell(row=i,column=j,value=i)
# wb.create_sheet("Second sheet")
# wb.save(filename=r"E:\aniket mate\Excel_file_handling\first.xlsx")





