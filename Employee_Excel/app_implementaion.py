
from app_classes import Address
import time

from openpyxl.styles import Alignment

from app_abstract_classes import EmployeeAbstraction
from app_excel_oper import FILE_PATH, get_workbook_sheet
from data import generate_employee,data_with_multiple_addresses
from app_classes import Employee





def align_cell(sheet):
    for i in range(1, sheet.max_row+1):
        for j in range(1, sheet.max_column+1):
            cell = sheet.cell(row=i,column=j)
            cell.alignment = Alignment(horizontal = "center",vertical = "center")




class EmployeeOperation():
    def write_headers(self):
        wb, sheet = get_workbook_sheet()
        # print(wb,sheet)
        sheet.merge_cells("E1:G1")
        headers = "ID   Name  Age   Salary  Address"
        splitted_headers= headers.split()
        # print(splitted_headers)
        c=1
        for i in splitted_headers:
            sheet.cell(row=1,column=c).value = i
            # sheet.cell.alignment = Alignment(horizontal = "center",vertical = "center")
            c+=1

        # for i in range(1, sheet.max_column):
        align_cell(sheet)
        


        address_headers = "Pincode City State"
        address = address_headers.split()
        sheet["E2"] = address[0]
        sheet["F2"] = address[1]
        sheet["G2"] = address[2]
        wb.save(FILE_PATH)
        return("Headers written successfully")

    
    # def write_data_to_excel(self):
    #     workbook, sheet = get_workbook_sheet()
    #     employee_list = generate_employee(1000)
    #     counter = 3
    #     for emp in employee_list:
    #         # print(emp)
    #         sheet.cell(row=counter  , column=1).value=emp.EmpID
    #         sheet.cell(row=counter  , column=2).value=emp.EmpName
    #         sheet.cell(row=counter  , column=3).value=emp.EmpAge
    #         sheet.cell(row=counter  , column=4).value=emp.EmpSalary
    #         sheet.cell(row=counter  , column=5).value=emp.EmpAddress.AddrPin
    #         sheet.cell(row=counter  , column=6).value=emp.EmpAddress.AddrCity
    #         sheet.cell(row=counter  , column=7).value=emp.EmpAddress.AddrState
    #         counter+=1
    #     align_cell(sheet)
        
    #     workbook.save(FILE_PATH)
    #     return("data inserted successfully")

        

    # def read_data_from_excel(self):
    #     workbook, sheet = get_workbook_sheet()
    #     fetched_list_emp =[]
    #     for i in range(3, sheet.max_row+1):
    #         EmpID = sheet.cell(row=i, column=1).value
    #         EmpName = sheet.cell(row=i, column=2).value
    #         EmpAge = sheet.cell(row=i, column=3).value
    #         EmpSalary= sheet.cell(row=i, column=4).value
    #         EmpPincode = sheet.cell(row=i, column=5).value
    #         EmpCity = sheet.cell(row=i, column=6).value
    #         EmpState = sheet.cell(row=i, column=7).value
    #         emp = Employee(eid=EmpID,ename=EmpName,eage=EmpAge,esal=EmpSalary,
    #                 eadr=[Address(pin= EmpPincode,city=EmpCity,state=EmpState)])      
    #         fetched_list_emp.append(emp)
    #     return(fetched_list_emp)


    # def write_data_to_excel(self):
    #     workbook, sheet = get_workbook_sheet()
    #     emp_list = data_with_multiple_addresses()
    #     # print(emp_list)
    #     counter = 3
    #     for emp in emp_list:
    #         sheet.cell(row=counter  , column=1).value=emp.EmpID
    #         sheet.cell(row=counter  , column=2).value=emp.EmpName
    #         sheet.cell(row=counter  , column=3).value=emp.EmpAge
    #         sheet.cell(row=counter  , column=4).value=emp.EmpSalary
    #         for adr in emp.EmpAddress:
    #             sheet.cell(row=counter  , column=5).value=adr.AddrPin
    #             sheet.cell(row=counter  , column=6).value=adr.AddrCity
    #             sheet.cell(row=counter  , column=7).value=adr.AddrState
    #             counter +=1
    #     workbook.save(FILE_PATH)
        

    def read_data_from_excel(self):
        workbook, sheet = get_workbook_sheet()
        fetched_list_emp =[]
        
        for i in range(3, sheet.max_row+1):
            EmpID = sheet.cell(row=i, column=1).value
            EmpName = sheet.cell(row=i, column=2).value
            EmpAge = sheet.cell(row=i, column=3).value
            EmpSalary= sheet.cell(row=i, column=4).value
            for j in range(3, sheet.max_row+1):
                EmpPincode = sheet.cell(row=j, column=5).value
                EmpCity = sheet.cell(row=j, column=6).value
                EmpState = sheet.cell(row=j, column=7).value
                emp = Employee(eid=EmpID,ename=EmpName,eage=EmpAge,esal=EmpSalary,
                            eadr=[Address(pin= EmpPincode,city=EmpCity,state=EmpState)])
            fetched_list_emp.append(emp)
        return fetched_list_emp


# workbook, sheet = get_workbook_sheet()

# EmpState = sheet.cell(row=4, column=6).value
# print(EmpState)

if __name__ == '__main__':

    emp_imp1 = EmployeeOperation()
#     t1 = time.time()
    # emp_imp1.write_headers()
    # emp_imp1.write_data_to_excel()
# #     t2 = time.time()
# #     print(f"Time taken:- {t2 - t1}
    # emp_imp1.read_data_from_excel()
    print(emp_imp1.read_data_from_excel())

"""
#----------------------------------------------OUTPUT-----------------------------------------------
{'EmpID': 101, 'EmpName': 'ABC', 'EmpAge': 23, 'EmpSalary': 45623, 'EmpAddress': [{'AddrPin': 789456, 'AddrCity': 'Panvel', 'AddrState': 'MH'}]},
{'EmpID': None, 'EmpName': None, 'EmpAge': None, 'EmpSalary': None, 'EmpAddress': [{'AddrPin': 789456, 'AddrCity': 'Panvel', 'AddrState': 'MH'}]},
{'EmpID': 102, 'EmpName': 'XYZ', 'EmpAge': 23, 'EmpSalary': 47896, 'EmpAddress': [{'AddrPin': 789456, 'AddrCity': 'Panvel', 'AddrState': 'MH'}]},
{'EmpID': 103, 'EmpName': 'PQR', 'EmpAge': 23, 'EmpSalary': 56563, 'EmpAddress': [{'AddrPin': 789456, 'AddrCity': 'Panvel', 'AddrState': 'MH'}]},
{'EmpID': None, 'EmpName': None, 'EmpAge': None, 'EmpSalary': None, 'EmpAddress': [{'AddrPin': 789456, 'AddrCity': 'Panvel', 'AddrState': 'MH'}]},
{'EmpID': None, 'EmpName': None, 'EmpAge': None, 'EmpSalary': None, 'EmpAddress': [{'AddrPin': 789456, 'AddrCity': 'Panvel', 'AddrState': 'MH'}]},
{'EmpID': 104, 'EmpName': 'STU', 'EmpAge': 23, 'EmpSalary': 45214, 'EmpAddress': [{'AddrPin': 789456, 'AddrCity': 'Panvel', 'AddrState': 'MH'}]},
{'EmpID': None, 'EmpName': None, 'EmpAge': None, 'EmpSalary': None, 'EmpAddress': [{'AddrPin': 789456, 'AddrCity': 'Panvel', 'AddrState': 'MH'}]}]
"""